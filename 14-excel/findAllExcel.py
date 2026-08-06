import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.utils import get_column_letter as gcl
from rich import print as rprint

search_re = re.compile(re.escape(sys.argv[1]), re.IGNORECASE)
found = defaultdict(list[str])

for excel_file in (x for x in Path.cwd().iterdir() if x.suffix == ".xlsx"):
    try:
        wb = openpyxl.load_workbook(excel_file)

        for sheetname in wb.sheetnames:
            sheet = wb[sheetname]
            for row_idx in range(1, sheet.max_row + 1):
                for col_idx in range(1, sheet.max_column + 1):
                    if search_re.search(str(sheet.cell(row_idx, col_idx).value)):
                        found[f"{excel_file.name} - {sheet.title}"].append(
                            f"{gcl(col_idx)}{row_idx}"
                        )
    except Exception as e:
        rprint(f"{e}")

if found:
    for k, v in found.items():
        rprint(f"[blue]{k}: [green]{v}\n")
else:
    rprint("[red]Nothing found!")
