import sys

import openpyxl
from openpyxl.utils import get_column_letter as gcl

wb = openpyxl.Workbook()
sheet = wb.active; assert sheet is not None
sheet.title = "Multiplication Table"

for offset in range(1, int(sys.argv[1]) + 1):
    sheet[f"A{offset + 1}"] = offset
    sheet[f"{gcl(offset + 1)}1"] = offset
for col in range(1, int(sys.argv[1]) + 1):
    for row in range(1, int(sys.argv[1]) + 1):
        sheet.cell(row + 1, col + 1).value = row * col  # type: ignore

wb.save("multiplicationTable.xlsx")
