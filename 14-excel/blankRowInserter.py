import sys

import openpyxl

N, M, name = int(sys.argv[1]), int(sys.argv[2]), sys.argv[3]
wb = openpyxl.load_workbook(name)
sheet = wb["Data"]
rows = sheet.max_row
cols = sheet.max_column + 1

for row_idx in range(rows, N - 1, -1):
    for col_idx in range(1, cols):
        sheet.cell(row_idx + M, col_idx).value = sheet.cell(row_idx, col_idx).value  # type:ignore

for row_idx in range(N, N + M):
    for col_idx in range(1, sheet.max_column + 1):
        sheet.cell(row_idx, col_idx).value = ""  # type: ignore

wb.save(name)
